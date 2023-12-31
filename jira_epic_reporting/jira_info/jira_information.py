import os, requests, sys, argparse
import openpyxl
from typing import List
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from colorama import init, Fore, Back, Style
from dotenv import load_dotenv
from datetime import datetime
from epic import Epic
from issue import Issue
from sprint import Sprint
from jirainfo import Jirainfo

start_time = datetime.now()
start_time_format = start_time.strftime("%m/%d/%Y, %H:%M:%S")
date_file_info = start_time.strftime("%Y_%m_%d")
create_date = start_time.strftime("%m/%d/%Y")

load_dotenv()
init() # Colorama   

jirakey = os.getenv("JIRA_API_KEY")
url_location = os.getenv("JIRA_REV_LOCATION")
url_search = os.getenv("JIRA_SEARCH")
url_board = os.getenv("JIRA_BOARD")
path_location = os.getenv("PATH_LOCATION")
project_label = os.getenv("PROJECT_LABEL")
jira_issue_link = os.getenv("JIRA_ISSUE_LINK")
project_figma_link = os.getenv("FIGMA_LINK")

main_serach = f"{url_location}/{url_search}"
board_info = f"{url_location}/{url_board}"
header = {"Authorization": "Basic " + jirakey}
baord_issues = f"{url_location}/{url_board}"
epics: List[Epic] = []  
temp_boards = {}
temp_sprints = {}

def terminal_update(message, data, bold):
    if bold:
        print(Back.GREEN + Fore.BLACK + Style.BRIGHT + f"  {message}: " + Back.BLUE + Fore.BLACK + Style.BRIGHT + f" {data} " + Style.RESET_ALL, end="\r")
    else:
        print(Fore.GREEN + Style.BRIGHT + f"  {message}: " + Fore.BLUE + Style.NORMAL + f" {data} " + Style.RESET_ALL, end="\r")

def terminal_busy(message, count):
    if count > 3:
        count = 0
    data = ""
    if count == 0:
        data = "|"
    elif count == 1:
        data = "/"
    elif count == 2:
        data = "-"
    elif count == 3:
        data = "\\"
    print(Fore.GREEN + Style.BRIGHT + f"  {message}: " + Fore.BLUE + Style.NORMAL + f" {data} " + Style.RESET_ALL, end="\r")

def get_boards(con_out):
    terminal_update("Retrieving Boards", " - ", False)
    # https://revlocaldev.atlassian.net/rest/agile/1.0/board
    response = requests.get(board_info, headers=header)
    if response.status_code == 200:
        data = response.json()
        for boarditem in data["values"]:
            temp_boards[boarditem["id"]] = f"{boarditem['name']}|{boarditem['type']}"
            # print(f"ID - {boarditem['id']} - Name - {boarditem['name']} - Type - {boarditem['type']}")
            # location
            #           projectId
            #           name
            #           projectKey
            #           projectTypeKey
            #           displayName
            #           projectNmae
        if con_out: 
            print(Fore.GREEN + f"Success! - Board Info {len(epics)}" + Style.RESET_ALL)
        else:
            if con_out:
                print(Fore.RED + "Failed - All Boards Info" + Style.RESET_ALL)

def get_sprints(con_out):
    # https://revlocaldev.atlassian.net/rest/agile/1.0/board/BOARD/sprint?maxResults=50&startAt=46
    count = 0
    for boarditem in temp_boards:
        start_location = 0
        count_boarditem_sprints = 0
        has_more_sprints = True
        while (has_more_sprints):
            terminal_busy("Retrieving Sprints", count)
            count += 1
            if count > 3:
                count = 0
            sprint_info = f"{url_location}/rest/agile/1.0/board/{boarditem}/sprint?maxResults=50&startAt={start_location}"
            response = requests.get(sprint_info, headers=header)
            if response.status_code == 200:
                data = response.json()
                for sprintitem in data["values"]:
                    #print(f"[{boarditem}-{count_boarditem_sprints}] ID - {sprintitem['id']} - Name - {sprintitem['name']} - State - {sprintitem['state']}")
                    temp_sprints[sprintitem["id"]] = f"{boarditem}|{count_boarditem_sprints}|{sprintitem['name']}|{sprintitem['state']}|{sprintitem['startDate']}|{sprintitem['endDate']}"
                    count_boarditem_sprints += 1
                    #"values": [
                    #    {
                    #        "id": 529,
                    #        "self": "https://revlocaldev.atlassian.net/rest/agile/1.0/sprint/529",
                    #        "state": "closed",
                    #        "name": "ARR - Sprint 139 (1/4-1/17)",
                    #        "startDate": "2023-01-04T16:54:31.502Z",
                    #        "endDate": "2023-01-17T22:00:00.000Z",
                    #        "completeDate": "2023-01-18T16:33:07.983Z",
                    #        "createdDate": "2023-01-03T16:33:18.087Z",
                    #        "originBoardId": 70,
                    #        "goal": "1. Implement custom block builder property list and add properties\n2. Implement custom block renderer \n3. Update property editor to pin most relevant block properties at the top"
                    #    },
                if len(data["values"]) == 50:
                    has_more_sprints = True
                    start_location += 50
                else:
                    has_more_sprints = False
            else:
                has_more_sprints = False
    if con_out:
        print(Fore.GREEN + f"Success! - Sprint Info {len(epics)}" + Style.RESET_ALL)
    else:
        if con_out:
            print(Fore.RED + "Failed - All Sprint Info" + Style.RESET_ALL)

# Retrieve all epics from main project label
# Get Sub labels to help break down the epics
def get_epics(label, con_out):
    terminal_update("Retrieving Epics", " - ", False)
    all_epics = main_serach + "'issuetype'='Epic' AND ('Status'='FUTURE' OR 'Status'='NEXT' OR 'Status'='Now') AND 'labels' in ('" + label + "')"
    response = requests.get(all_epics, headers=header)
    if response.status_code == 200:
        data = response.json()
        for epicitem in data["issues"]:
            epic_add = Epic(epicitem["id"], epicitem["key"], epicitem["fields"]["summary"], epicitem["fields"]["created"])
            epic_add.set_team(epicitem["fields"]["project"]["name"])
            epic_add.set_estimate(epicitem["fields"]["customfield_10032"])
            epic_add.description = epicitem["fields"]["description"]
            for label in epicitem["fields"]["labels"]:
                if (label != project_label):
                    epic_add.add_sub_label(label)
            epics.append(epic_add)
        if con_out:
            print(Fore.GREEN + f"Success! - All Epics {len(epics)}" + Style.RESET_ALL)
    else:
        if con_out:
            print(Fore.RED + "Failed - All Epics" + Style.RESET_ALL)

# Retrieve all the issues attached to the epics
# JSON Issue
# issues
#      id - 12345                                           id
#      key - ARR-2392                                       key
#      fields
#            summary - "Create a new project report"        summary
#            customfield_10032 - 5 Points                   size  
#            customfield_10010 - Sprint[]                   sprint[]
#            status
#                  name - "In Progress"                     status
#            priority
#                  name - "Medium"                          priority
#            issuetype
#                  name - "Story"                           issuetype
#            project
#                  name - "Agile RevSite Raider$"           project_name
#                  key - "ARR"                              project_key
#            assignee
#                  displayName - "John Doe"                 assignee_displayName
#            created - "2021-03-01T15:00:00.000-0400        created
#            updated - "2021-03-01T15:00:00.000-0400        updated
#            description - "This is a description"          description

def get_issues():
    issues_with_points = 0
    issues_points = 0
    issues_with_no_points = 0
    count_epics = (len(epics)+1)
    count_epics_current = 1
    for epicitem in epics:
        terminal_update("Retrieving Issues on Epics", f"{count_epics_current}/{count_epics}", False)
        count_epics_current += 1
        epic_issues = main_serach + "'Epic Link'='" + epicitem.key + "' and STATUS != Cancelled"
        response = requests.get(epic_issues, headers=header)
        if response.status_code == 200:
            data = response.json()
            for issue in data["issues"]:
                assigned_points = issue["fields"]["customfield_10032"]
                issue_add = Issue(issue["id"], issue["key"], issue["fields"]["summary"], assigned_points)
                issue_add.set_status(issue["fields"]["status"]["name"])
                issue_add.set_priority(issue["fields"]["priority"]["name"])
                issue_add.set_issuetype(issue["fields"]["issuetype"]["name"])
                issue_add.set_project_name(issue["fields"]["project"]["name"])
                issue_add.set_project_key(issue["fields"]["project"]["key"])
                if issue["fields"]["assignee"] is not None:
                    if "displayName" in issue["fields"]["assignee"]:
                        issue_add.set_assignee_displayName(issue["fields"]["assignee"]["displayName"])
                issue_add.set_created(issue["fields"]["created"])
                issue_add.set_updated(issue["fields"]["updated"])
                issue_add.set_description(issue["fields"]["description"])
                
                if assigned_points == None:
                    issues_with_no_points += 1
                else:
                    issues_with_points += 1
                    try: 
                        issues_points += assigned_points
                    except:
                        pass
                try:
                    for item in issue["fields"]["customfield_10010"]:
                        add_sprint = Sprint(item["id"], item["name"], item["boardId"], item["state"])
                        if "completeDate" in item:
                            add_sprint.set_completeDate(item["completeDate"])
                        issue_add.add_sprint(add_sprint)
                except:
                    pass
                epicitem.add_issue(issue_add)
                epicitem.set_issues_with_points(issues_with_points)
                epicitem.set_issues_points(issues_points)
                epicitem.set_issues_with_no_points(issues_with_no_points)
    
# Console Output Information
def output_console():
    for epicitem in epics:
        epicitem.print_Epic()
        for issueitem in epicitem.issues:
            issueitem.print_Issue()
            for sprintitem in issueitem.sprint:
                sprintitem.print_Sprint()

# Create new workbook
# --- Summary Tab ---
# A1           B1       C1            D1            E1            F1            G1            H1
# Label                 Created
#                      
# [Epics]              [Issues Overall]
# Count                 Count
# Total Estimate        Total Estimate
# With Estimates        With Estimates
# Percent               Percent
# Average Estimate      Average Estimate
# Max Estimate          Max Estimate              
# Min Estimate          Min Estimate
# Issues in Epic
# [Count]
# Total Estimate
# With Estimates
# Percent
# Average Estimate
# Max Estimate
# Min Estimate
#
#
#
# --- Epic Tab ---
# Key(link) | Summary | Team | Estimate | Issues with Points | Issues with No Points | Issues Points | Sub Labels
#
# --- Issue Tab ---
# Define
#
def create_excel(label):
    terminal_update("Creating Excel Document", " - ", False)
    workbook = openpyxl.Workbook()
    worksheet_summary = workbook.active
    worksheet_summary.title = "Summary"
    worksheet_epics = workbook.create_sheet("Epics")
    worksheet_issues = workbook.create_sheet("Issues")

    # Create the Summary Tab
    Epic.excel_worksheet_summary(worksheet_summary, epics, label, create_date)

    # Create the Epic Tab
    Epic.excel_worksheet_create(worksheet_epics, epics, jira_issue_link, project_figma_link,label)

    # Create the Issue Tab
    Issue.excel_worksheet_create(worksheet_issues, epics, jira_issue_link)

    return workbook

def create_excel_jira_summary():
    workbook = openpyxl.Workbook()
    worksheet_boards = workbook.active
    worksheet_boards.title = "Boards"
    worksheet_sprints = workbook.create_sheet("Sprints")
    Jirainfo.excel_boards(worksheet_boards, temp_boards)
    Jirainfo.excel_sprints(worksheet_sprints, temp_sprints)
    save_file_info(path_location, "Jira Info", workbook)

def save_file_info(path, filename, wb):
    # Handle Directory
    if os.path.exists(path):
        saveexcelfile = path + date_file_info + " " + filename + ".xlsx"
    else:
        os.makedirs(path)

    # Check For File Existence - Delete if exists
    if os.path.exists(saveexcelfile):
        os.remove(saveexcelfile)

    # Save Workbook
    wb.save(saveexcelfile)

def save_file(path, filename, wb):
    # Handle Directory
    if os.path.exists(path):
        saveexcelfile = path + date_file_info + " Project " + filename + " Details.xlsx"
    else:
        os.makedirs(path)

    # Check For File Existence - Delete if exists
    if os.path.exists(saveexcelfile):
        os.remove(saveexcelfile)

    # Save Workbook
    wb.save(saveexcelfile)

def main(args):
    if args.label:
        project_label = args.label

    con_out = False
    if args.console:
        con_out = True
    
    # Get Boards
    get_boards(con_out)
    # Get Sprints
    get_sprints(con_out)
    # Get People

    # Get Issues People have be assigned

    create_excel_jira_summary()

    #get_epics(project_label, con_out)
    #get_issues()
    #wb = create_excel(project_label)
    #save_file(path_location,project_label,wb)

    #if con_out:
    #    output_console()  

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Create Excel Sheet for Project Reporting")
    parser.add_argument("-l", "--label", help="Label for the project")
    parser.add_argument("-c", "--console", help="Enable Console Output", action="store_true")
    args = parser.parse_args()
    main(args)
