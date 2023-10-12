from typing import List
from issue import Issue
from colorama import init, Fore, Back, Style
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import numbers
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

class Jirainfo:
    def excel_boards(ws, boards):
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 25
        
        table = Table(displayName="TableBoards", ref="A1:C" + str(len(boards) + 1))
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style

        ws.append(["Board ID", "Board Name", "Board Type"])
        
        for board in boards:
            name, type = boards[board].split("|")
            ws.append([board,name,type])

        ws.add_table(table)

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[0] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row in ws[2:ws.max_row]:  # Include The Header
            cell = row[1] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[3:ws.max_row]:  # Include The Header
            cell = row[1] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def excel_sprints(ws, sprints):
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 45
        ws.column_dimensions["E"].width = 25

        ws.append(["Sprint ID", "Board ID", "Number", "Name", "State"])
        
        table = Table(displayName="TableSprints", ref="A1:E" + str(len(sprints) + 1))
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style

        for sprint in sprints:
            board_id, sprint_number, sprint_name, sprint_state = sprints[sprint].split("|")
            ws.append([sprint, board_id, sprint_number, sprint_name, sprint_state])

        ws.add_table(table)
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[0] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[0] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[2] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[3] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[4] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")
