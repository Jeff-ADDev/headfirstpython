from typing import List
from colorama import init, Fore, Back, Style
from sprint import Sprint
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Retrieve all the issues attached to the epics
# JSON Issue
# issues
#      id - 12345                                           id
#      key - ARR-2392                                       key
#      fields
#            summary - "Create a new project report"        summary
#            customfield_10032 - 5 Points                   size  
#            customfield_10010 - Sprint[]                   sprint[]
#            status
#                  name - "In Progress"                     status
#            priority
#                  name - "Medium"                          priority
#            issuetype
#                  name - "Story"                           issuetype
#            project
#                  name - "Agile RevSite Raider$"           project_name
#                  key - "ARR"                              project_key
#            assignee
#                  displayName - "John Doe"                 assignee_displayName
#            created - "2021-03-01T15:00:00.000-0400        created
#            updated - "2021-03-01T15:00:00.000-0400        updated
#            description - "This is a description"          description

class Issue:
    def __init__(self, id, key, summary, size):
        self.id = id
        self.key = key
        self.summary = summary
        self.size = size
        self.sprint: List[Sprint] = []
        self.status = ""
        self.priority = ""
        self.issuetype = ""
        self.project_name = ""
        self.project_key = ""
        self.assignee_displayName = ""
        self.created = ""
        self.updated = ""
        self.description = ""
    
    def set_status(self, status):
        self.status = status
    
    def set_priority(self, priority):
        self.priority = priority
    
    def set_issuetype(self, issuetype):
        self.issuetype = issuetype
    
    def set_project_name(self, project_name):
        self.project_name = project_name
    
    def set_project_key(self, project_key):
        self.project_key = project_key

    def set_assignee_displayName(self, assignee_displayName):
        self.assignee_displayName = assignee_displayName

    def set_created(self, created):
        self.created = datetime.strptime(created, "%Y-%m-%dT%H:%M:%S.%f%z")
    
    def set_updated(self, updated):
        self.updated = datetime.strptime(updated, "%Y-%m-%dT%H:%M:%S.%f%z")

    def set_description(self, description):
        self.description = description

    def add_sprint(self, sprint):
        self.sprint.append(sprint)

    def set_sprint_name(self, sprint_name):
        self.sprint_name = sprint_name

    def set_sprint_state(self, sprint_state):
        self.sprint_state = sprint_state

    def set_boardID(self, boardID):
        self.boardID = boardID

    def set_completeDate(self, completeDate):
        self.completeDate = completeDate

    def print_Issue(self):
        print(Fore.CYAN + Style.BRIGHT + 
              "  Issue-" + self.key + ": " + Fore.CYAN + Style.NORMAL + self.summary 
              + Fore.BLUE + Style.BRIGHT + " (" + str(self.size) + ")" + Style.RESET_ALL)
        
    def excel_worksheet_create(ws, epics, jira_issue_link):
        def set_date(value):
            if value != "":
                return value.strftime("%m/%d/%Y")
            else:
                return ""

        # Start Building Issues Tab
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 40
        ws.column_dimensions["I"].width = 20
        ws.column_dimensions["J"].width = 20
        ws.column_dimensions["K"].width = 20
        ws.column_dimensions["L"].width = 20
        ws.column_dimensions["M"].width = 30
        ws.column_dimensions["N"].width = 30
        ws.column_dimensions["O"].width = 30
        ws.column_dimensions["P"].width = 30



        all_issues = 0
        for epicitem in epics:
            all_issues += (len(epicitem.issues))
        table_issues = Table(displayName="TableIssues", ref="A1:P" + str(all_issues + 1))
        style_issues = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table_issues.tableStyleInfo = style_issues

        ws.append(["Issue", "Summary", "Team", "Size", 
                   "Status", "Priority", "Type", 
                   "Project", "Assignee", "Created", 
                   "Updated", "Epic",
                   "Sprint", "Sprint 2", "Sprint 3", "Sprint 4"])
        for epicitem in epics:
            for issueitem in epicitem.issues:
                sprint_count = 0
                sprint1 = ""
                sprint2 = ""
                sprint3 = ""
                sprint4 = ""
                for sprintitem in issueitem.sprint:
                    if sprint_count == 0:
                        sprint1 = sprintitem.name
                    elif sprint_count == 1:
                        sprint2 = sprintitem.name
                    elif sprint_count == 2:
                        sprint3 = sprintitem.name
                    elif sprint_count == 3:
                        sprint4 = sprintitem.name
                    sprint_count += 1

                ws.append([issueitem.key, issueitem.summary, epicitem.team, issueitem.size
                            , issueitem.status, issueitem.priority, issueitem.issuetype
                            , issueitem.project_name, issueitem.assignee_displayName, set_date(issueitem.created)
                            , set_date(issueitem.updated), epicitem.key,
                            sprint1, sprint2, sprint3, sprint4])

        ws.add_table(table_issues)

        for row in ws[2:ws.max_row]:  # Exclude The Header
            cell = row[0] # zeor based index
            value_use = cell.value
            cell.hyperlink = f"{jira_issue_link}{value_use}"
            cell.value = value_use
            cell.style = "Hyperlink"
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[0] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[2] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[3] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[4] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[5] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[6] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[7] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[8] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[9] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[10] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[2:ws.max_row]:  # Exclude The Header
            cell = row[11] # zeor based index
            value_use = cell.value
            cell.hyperlink = f"{jira_issue_link}{value_use}"
            cell.value = value_use
            cell.style = "Hyperlink"
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[11] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")