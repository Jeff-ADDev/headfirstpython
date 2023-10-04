import os, requests
import openpyxl
from openpyxl.utils import get_column_letter
from colorama import init, Fore, Back, Style
from dotenv import load_dotenv
from datetime import datetime

start_time = datetime.now()
start_time_format = start_time.strftime("%m/%d/%Y, %H:%M:%S")

load_dotenv()
init() # Colorama   

jirakey = os.getenv("JIRA_API_KEY")
url_location = os.getenv("JIRA_REV_LOCATION")
url_search = os.getenv("JIRA_SEARCH")
sheets_location = os.getenv("SHEETS_LOCATION")


main_serach = f"https://{url_location}/{url_search}"
header = {"Authorization": "Basic " + jirakey}

# Retrieve all epics
# Teamm, Status, Labels
# issuetype = Epic AND project = "Agile RevSite Raider$" AND (Status = 'FUTURE' OR Status = 'NEXT' OR Status = 'Now') AND labels in (ReviewMarketing)
all_epics = main_serach + "'issuetype'='Epic' AND 'project'='Agile RevSite Raider$' AND ('Status'='FUTURE' OR 'Status'='NEXT' OR 'Status'='Now') AND 'labels' in ('ReviewMarketing')"
response = requests.get(all_epics, headers=header)
if response.status_code == 200:
    print(Fore.GREEN + "Success! - All Epics" + Style.RESET_ALL)
    data = response.json()
    for issue in data["issues"]:
        print(issue["fields"]["summary"])
else:
    print(Fore.RED + "Failed!")

epic_info = main_serach + "'issue'='ARR-2392'"
response = requests.get(epic_info, headers=header)
if response.status_code == 200:
    print(Fore.GREEN + "Success! - Epic Item" + Style.RESET_ALL)
    data = response.json()
    for issue in data["issues"]:
        print(issue["fields"]["summary"])
else:
    print(Fore.RED + "Failed!")

epic_issues = main_serach + "'Epic Link'='ARR-2392' and STATUS != Cancelled"
response = requests.get(epic_issues, headers=header)
if response.status_code == 200:
    print(Fore.GREEN + "Success! - Epic Issues" + Style.RESET_ALL)
    data = response.json()
    for issue in data["issues"]:
        print(Fore.YELLOW + Style.BRIGHT 
              + issue["key"] + " - "
              + issue["fields"]["summary"] + Style.RESET_ALL)
        
        try:
            for item in issue["fields"]["customfield_10010"]:
                cf_10010_name = item["name"]
                cf_10010_state = item["state"]
        except:
                cf_10010_name = ""
                cf_10010_state = ""
        
        print(Fore.LIGHTYELLOW_EX + Style.DIM
              + issue["fields"]["issuetype"]["name"] 
              + " : " + str(issue["fields"]["customfield_10032"])
              + " : " + issue["fields"]["project"]["name"] 
              + " : " + cf_10010_name
              + " : " + cf_10010_state
              + Style.RESET_ALL)

else:
    print(Fore.RED + "Failed!")

# Create new workbook
workbook = openpyxl.Workbook()
worksheet_epics = workbook.active
worksheet_epics.title = "Epics"

worksheet_links = workbook.create_sheet("Links")

worksheet_logging = workbook.create_sheet("Logging")

worksheet_epics["A1"] = "Epic Name"

worksheet_links["A1"].hyperlink = "https://www.figma.com/file/7YgLsbg9xLvPJmuYAZvluI/Site-Builder-Roadmap?type=whiteboard&node-id=0%3A1&t=IHX8Dq3vHGjn3th1-1Z3jg"
worksheet_links['A1'].value = 'Figma Plan'
worksheet_links['A1'].style = "Hyperlink"

worksheet_logging.column_dimensions["A"].width = 40
worksheet_logging["A1"] = start_time_format

# Handle Directory
if os.path.exists(sheets_location):
    saveexcelfile = sheets_location + "epic_reporting.xlsx"
else:
    os.makedirs(sheets_location)

# Check For File Existence - Delete if exists
if os.path.exists(saveexcelfile):
    os.remove(saveexcelfile)

workbook.save(saveexcelfile)
