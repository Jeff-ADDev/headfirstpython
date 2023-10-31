import os, argparse
from typing import List
from colorama import init, Fore, Back, Style
from dotenv import load_dotenv
from datetime import datetime, timedelta
import console_util
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, numbers, PatternFill, Border, Side, colors, GradientFill
from openpyxl.styles.fills import Stop
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice, GradientFillProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors
from copy import deepcopy
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor
from objects.calendar_sprint import CalendarSprint

from openpyxl.chart import (
    LineChart,
    BarChart,
    ScatterChart,
    Reference,
    Series,
)



init() # Colorama   

def bar_chart():
    """
        https://openpyxl.readthedocs.io/en/latest/charts/bar.html
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Boards"    

    ws.column_dimensions["A"].width = 18 
    ws.column_dimensions["B"].width = 18 
    ws.column_dimensions["C"].width = 18 

    rows = [
        ['Sprint', 'Stories Added', 'Other'],
        [160, 5, 6],
        [161, 8, 7],
        [162, 12,15],
        [163, 15, 20],
        [164, 25, 22],
        [165, 2, 5],
        [166, 2, 7],
        [166, 0, 2],
    ]

    for row in rows:
        ws.append(row)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Stories Added Per Sprint"
    chart.style = 10
    chart.x_axis.title = 'Sprint'
    chart.y_axis.title = 'Count'

    max_row_set = ws.max_row

    data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=2)
    cats = Reference(ws, min_col=1, min_row=2, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4 

    #xvalues = Reference(ws, min_col=1, min_row=2, max_row=max_row_set) # Size Column
    #for i in range(2, 4): # Column (2, 5) 2,3,4: B and C and D
    #    values = Reference(ws, min_col=i, min_row=1, max_row=max_row_set)
    #    series = Series(values, xvalues, title_from_data=True)
    #    chart.series.append(series) 

    ws.add_chart(chart, "F2")

    console_util.save_excel_file("./", "Python Excel Example Chart.xlsx", wb)

def simple_chart():
    wb = Workbook()
    ws = wb.active
    ws.title = "Boards"

    ws.column_dimensions["A"].width = 18 
    ws.column_dimensions["B"].width = 18 
    ws.column_dimensions["C"].width = 18 
    ws.column_dimensions["D"].width = 18 

    rows = [
        ['Sprint', 'Total Points', 'Completed Points','Estimated Trajectory'],
        [160, 80, 10, 0],
        [161, 140, 25, "=NA()"],
        [162, 210, 45, "=NA()"],
        [163, 230, 80, "=NA()"],
        [164, 235, 100, "=NA()"],
        [165, 235, "", "=NA()"],
        [166, 235, "", "=NA()"],
        [166, 235, "", 235],
    ]

    for row in rows:
        ws.append(row)



    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[0] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[1] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[2] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[3] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    chart = ScatterChart()
    chart.title = "Project Burn Up Chart"
    chart.style = 6
    chart.x_axis.title = 'Size'
    chart.y_axis.title = 'Percentage'

    max_row_set = ws.max_row

    xvalues = Reference(ws, min_col=1, min_row=2, max_row=max_row_set) # Size Column

    for i in range(2, 5): # Column (2, 5) 2,3,4: B and C and D
        values = Reference(ws, min_col=i, min_row=1, max_row=max_row_set)
        series = Series(values, xvalues, title_from_data=True)
        chart.series.append(series) 




    # Chart Line Properties
    series2 = chart.series[2]
    line_props = LineProperties(solidFill="FF0000", prstDash="dot")
    series2.graphicalProperties.line = line_props

    series0 = chart.series[0]
    line_props0 = LineProperties(solidFill="04A24E", prstDash="solid")
    series0.graphicalProperties.line = line_props0
    chart2 = deepcopy(chart)

    # How to change the background color of a chart
    fill_properties = GradientFillProperties()
    first_color = "9CFDC9"
    second_color = "04A24E"
    fill = GradientFill(type='linear',
                 degree=90,
                 stop=(Stop(second_color, 0), Stop(first_color, 1)))
    fill_properties.gradientFill = GradientFill()
    chart.graphical_properties = GraphicalProperties()
    chart.graphical_properties.gradFill = fill_properties

    chart.graphical_properties.line.solidFill = "FF0000"
    chart.graphical_properties.line.prstDash = "dash"

    ws.add_chart(chart, "F2")
    ws.add_chart(chart2, "F18")
    anchor = TwoCellAnchor()
    anchor._from.col = 5 # F
    anchor._from.row = 18 # row 19, using 0-based indexing
    anchor.to.col = 20 # 
    anchor.to.row = 48 # row 
    chart2.anchor = anchor
    

    console_util.save_excel_file("./", "Python Excel Example Chart.xlsx", wb)


def main(args):
    start_time = datetime.now()
    start_time_format = start_time.strftime("%m/%d/%Y, %H:%M:%S")
    date_file_info = start_time.strftime("%Y %m %d")
    create_date = start_time.strftime("%m/%d/%Y")

    #simple_chart()
    bar_chart()

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Create Sampe Excel Sheet for Burnup Chart")
    parser.add_argument("-c", "--console", help="Enable Console Output", action="store_true")
    args = parser.parse_args()
    main(args)