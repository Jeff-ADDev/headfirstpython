import os, requests, sys
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from colorama import init, Fore, Back, Style
from dotenv import load_dotenv
from datetime import datetime
import epic

start_time = datetime.now()
start_time_format = start_time.strftime("%m/%d/%Y, %H:%M:%S")

load_dotenv()
init() # Colorama   

jirakey = os.getenv("JIRA_API_KEY")
url_location = os.getenv("JIRA_REV_LOCATION")
url_search = os.getenv("JIRA_SEARCH")
url_board = os.getenv("JIRA_BOARD")
sheets_location = os.getenv("SHEETS_LOCATION")


main_serach = f"{url_location}/{url_search}"
header = {"Authorization": "Basic " + jirakey}
baord_issues = f"{url_location}/{url_board}"

# 70 is board ID for RevSite Raiders
raider_baord_issues = baord_issues + "70/sprint?maxResults=50&startAt=50"
response = requests.get(raider_baord_issues, headers=header)
if response.status_code == 200:
    print(Fore.GREEN + "Success! - Raider Board Issues" + Style.RESET_ALL)
    data = response.json()
    for sprint in data["values"]:
        date_obj = datetime.strptime(sprint["startDate"], "%Y-%m-%dT%H:%M:%S.%fZ")
        print(str(date_obj.month) + "/" + str(date_obj.day) + "/" + str(date_obj.year))
        print(sprint["name"])
else:
    print(Fore.RED + "Failed - Raider Board Issues" + Style.RESET_ALL)

epics = []
# Retrieve all epics
# Teamm, Status, Labels
# issuetype = Epic AND project = "Agile RevSite Raider$" AND (Status = 'FUTURE' OR Status = 'NEXT' OR Status = 'Now') AND labels in (ReviewMarketing)
all_epics = main_serach + "'issuetype'='Epic' AND 'project'='Agile RevSite Raider$' AND ('Status'='FUTURE' OR 'Status'='NEXT' OR 'Status'='Now') AND 'labels' in ('ReviewMarketing')"
response = requests.get(all_epics, headers=header)
if response.status_code == 200:
    data = response.json()
    for epicitem in data["issues"]:
        print(epicitem["fields"]["summary"])
        epics.append(epic.Epic(epicitem["id"], epicitem["fields"]["summary"], epicitem["fields"]["created"]))
    print(Fore.GREEN + f"Success! - All Epics {len(epics)}" + Style.RESET_ALL)
else:
    print(Fore.RED + "Failed!")

epic_info = main_serach + "'issue'='ARR-2392'"
response = requests.get(epic_info, headers=header)
if response.status_code == 200:
    print(Fore.GREEN + "Success! - Epic Item" + Style.RESET_ALL)
    data = response.json()
    for epic in data["issues"]:
        print(epic["fields"]["summary"])
else:
    print(Fore.RED + "Failed!")

epic_issues = main_serach + "'Epic Link'='ARR-2392' and STATUS != Cancelled"
response = requests.get(epic_issues, headers=header)
if response.status_code == 200:
    print(Fore.GREEN + "Success! - Epic Issues" + Style.RESET_ALL)
    data = response.json()
    for epic in data["issues"]:
        print(Fore.YELLOW + Style.BRIGHT 
              + epic["key"] + " - "
              + epic["fields"]["summary"] + Style.RESET_ALL)
        
        try:
            for item in epic["fields"]["customfield_10010"]:
                cf_10010_name = item["name"]
                cf_10010_state = item["state"]
        except:
                cf_10010_name = ""
                cf_10010_state = ""
        
        print(Fore.LIGHTYELLOW_EX + Style.DIM
              + epic["fields"]["issuetype"]["name"] 
              + " : " + str(epic["fields"]["customfield_10032"])
              + " : " + epic["fields"]["project"]["name"] 
              + " : " + cf_10010_name
              + " : " + cf_10010_state
              + Style.RESET_ALL)

else:
    print(Fore.RED + "Failed!")

# Create new workbook
if (False):
    workbook = openpyxl.Workbook()
    worksheet_epics = workbook.active
    worksheet_epics.title = "Epics"

    worksheet_links = workbook.create_sheet("Links")

    worksheet_logging = workbook.create_sheet("Logging")

    worksheet_epics["A1"] = "Epic Name"

    worksheet_links["A1"].hyperlink = "https://www.figma.com/file/7YgLsbg9xLvPJmuYAZvluI/Site-Builder-Roadmap?type=whiteboard&node-id=0%3A1&t=IHX8Dq3vHGjn3th1-1Z3jg"
    worksheet_links['A1'].value = 'Figma Plan'
    worksheet_links['A1'].style = "Hyperlink"

    worksheet_logging.column_dimensions["A"].width = 40
    worksheet_logging["A1"] = start_time_format

    # Handle Directory
    if os.path.exists(sheets_location):
        saveexcelfile = sheets_location + "epic_reporting.xlsx"
    else:
        os.makedirs(sheets_location)

    # Check For File Existence - Delete if exists
    if os.path.exists(saveexcelfile):
        os.remove(saveexcelfile)

    workbook.save(saveexcelfile)

# Open Workbook
# Loop through rows and columns
if (False):
    workbook = None
    loadexcelfile = sheets_location + "epic_reporting_open.xlsx"
    if os.path.exists(loadexcelfile):
        workbook = load_workbook(loadexcelfile)
    else:
        print(Fore.RED + Style.BRIGHT + "Failed! - Can Not Open " + loadexcelfile + Style.RESET_ALL)
        sys.exit()
    
    worksheet_logging = workbook["Logging"]

    for cell in worksheet_logging["A"]:
        print(Fore.GREEN + "Cell Location " + cell.coordinate + Style.RESET_ALL)
        next_cell = cell.offset(row=1)
        #if cell.value is None:
        #    cell.value = start_time_format
        #    print(Fore.GREEN + "Setting Cell Value " + cell.coordinate + Style.RESET_ALL)
        #    break
    next_cell.value = start_time_format
    
    # Can itterate through rowa
    #for row in worksheet_logging.iter_rows(min_row=1, max_row=1, min_col=1, max_col=13):
    #    for cell in row:
    #        print(cell.value)

    workbook.save(loadexcelfile)
    print(Fore.GREEN + "Success! - Workbook Opened" + Style.RESET_ALL)

# Create a new workbook and table
if (True):
    workbook = openpyxl.Workbook()
    worksheet_epics = workbook.active
    worksheet_epics.title = "Epics"

    table = Table(displayName="TableTest", ref="A1:D5")

    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    table.tableStyleInfo = style

    data = [
        ["Name", "Age", "City", "Country"],
        ["John", 30, "New York", "USA"],
        ["Jane", 25, "Los Angeles", "USA"],
        ["Bob", 40, "London", "UK"],
        ["Alice", 35, "Paris", "France"]
    ]
    
    for row_data in data:
        worksheet_epics.append(row_data)

    worksheet_epics.add_table(table)

    # Handle Directory
    if os.path.exists(sheets_location):
        saveexcelfile = sheets_location + "epic_reporting_table.xlsx"
    else:
        os.makedirs(sheets_location)

    # Check For File Existence - Delete if exists
    if os.path.exists(saveexcelfile):
        os.remove(saveexcelfile)

    workbook.save(saveexcelfile)