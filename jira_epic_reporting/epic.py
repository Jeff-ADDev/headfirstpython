from typing import List
from issue import Issue
from colorama import init, Fore, Back, Style
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

class Epic:
    def __init__(self, id, key, summary, create_date):
        self.id = id
        self.key = key
        self.summary = summary
        self.create_date = datetime.strptime(create_date, "%Y-%m-%dT%H:%M:%S.%f%z")
        self.issues: List[Issue] = []
        self.sub_labels = []
        self.team = ""
        self.estimate = 0
        self.issues_with_points = 0
        self.issues_points = 0
        self.issues_with_no_points = 0

    def add_issue(self, issue):
        self.issues.append(issue)   

    def add_sub_label(self, sub_label):
        self.sub_labels.append(sub_label)
    
    def set_team(self, team):
        self.team = team
    
    def set_estimate(self, estimate):
        self.estimate = estimate

    def set_issues_with_points(self, issues_with_points):
        self.issues_with_points = issues_with_points
    
    def set_issues_points(self, issues_points):
        self.issues_points = issues_points
    
    def set_issues_with_no_points(self, issues_with_no_points):
        self.issues_with_no_points = issues_with_no_points

    def get_sublevles(self):
        sub_label_print = ""
        count_label = 0
        for sub_label in self.sub_labels:
            if (count_label == 0):
                sub_label_print += sub_label
                count_label += 1
            else:
                sub_label_print += sub_label + ", "
        return sub_label_print

    def print_Epic(self):
        print(Fore.YELLOW + Style.BRIGHT + 
              "Epic-" + str(self.key) + ": " + Fore.LIGHTYELLOW_EX + Style.NORMAL + self.summary + Fore.WHITE + 
              " Created: " + str(self.create_date.month) + "/" + str(self.create_date.day) + "/" + str(self.create_date.year) +
              "\n    " + Fore.BLUE + Style.BRIGHT + self.team + Fore.RED + Style.NORMAL + " Estimate: " + str(self.estimate) +
              "\n    " + Fore.YELLOW + Style.NORMAL + f"{self.issues_with_points} issues have points and {self.issues_with_no_points} don't. {self.issues_points} points total." +
              "\n    " + Fore.MAGENTA + " Sub Labels: " + Fore.WHITE + str(self.sub_labels) + Style.RESET_ALL)
        

    def excel_worksheet_create(ws, epics, jira_issue_link, project_figma_link):
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 40

        table = Table(displayName="TableEpics", ref="A1:H" + str(len(epics) + 1))
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style

        # Populate data
        ws.append(["Epic", "Summary", "Team", "Estimate", "Issues w Points", "Issues w No Points ", "Issues Total Points", "Sub Labels"])
        for epicitem in epics:
            sub_labels = epicitem.get_sublevles()
            ws.append([epicitem.key, epicitem.summary, epicitem.team, epicitem.estimate, epicitem.issues_with_points, epicitem.issues_with_no_points, epicitem.issues_points, sub_labels])    

        ws.add_table(table)

        # Format Data
        for row in ws[2:ws.max_row]:  # Exclude The Header
            cell = row[0] # zeor based index
            value_use = cell.value
            cell.hyperlink = f"{jira_issue_link}{value_use}"
            cell.value = value_use
            cell.style = "Hyperlink"
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[0] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[2:ws.max_row]:  # skip the header
            cell = row[1] # zeor based index
            cell.alignment = Alignment(wrap_text=True)
            cell.number_format = "text"

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[2] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[3] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[4] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[5] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[6] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add Figma Plan Link to the bottom of the Epics Sheet
        ws["A" + str(len(epics) + 3)].hyperlink = project_figma_link
        ws["A" + str(len(epics) + 3)].value = "Figma Plan"
        ws["A" + str(len(epics) + 3)].style = "Hyperlink"
