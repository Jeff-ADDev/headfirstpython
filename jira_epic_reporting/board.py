from typing import List
from objects.issue import Issue
from colorama import init, Fore, Back, Style
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import numbers
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

#            "id": 54,
#            "self": "https://revlocaldev.atlassian.net/rest/agile/1.0/board/54",
#            "name": "RevLocal Epic Detail",
#           "type": "kanban",
#            "location": {
#               "projectId": 10001,
#               "displayName": "Salesforce Development (SFD)",
#                "projectName": "Salesforce Development",
#                "projectKey": "SFD",
#                "projectTypeKey": "software",
#                "avatarURI": "https://revlocaldev.atlassian.net/rest/api/2/universal_avatar/view/type/project/avatar/10486?size=small",
#               "name": "Salesforce Development (SFD)"
#            }
 
class Board:
    def __init__(self, id, key, summary, create_date):
        self.id = id
        self.key = key
        self.summary = summary
        self.create_date = datetime.strptime(create_date, "%Y-%m-%dT%H:%M:%S.%f%z")
        self.issues: List[Issue] = []
        self.sub_labels = []
        self.team = ""
        self.estimate = 0
        self.issues_with_points = 0
        self.issues_points = 0
        self.issues_with_no_points = 0

    def add_issue(self, issue):
        self.issues.append(issue)   

    def add_sub_label(self, sub_label):
        self.sub_labels.append(sub_label)
    
    def set_team(self, team):
        self.team = team
    
    def set_estimate(self, estimate):
        self.estimate = estimate

    def set_issues_with_points(self, issues_with_points):
        self.issues_with_points = issues_with_points
    
    def set_issues_points(self, issues_points):
        self.issues_points = issues_points
    
    def set_issues_with_no_points(self, issues_with_no_points):
        self.issues_with_no_points = issues_with_no_points

    def get_sublevles(self):
        sub_label_print = ""
        count_label = 0
        for sub_label in self.sub_labels:
            if (count_label == 0):
                sub_label_print += sub_label
                count_label += 1
            else:
                sub_label_print += sub_label + ", "
        return sub_label_print

    def print_Epic(self):
        print(Fore.YELLOW + Style.BRIGHT + 
              "Epic-" + str(self.key) + ": " + Fore.LIGHTYELLOW_EX + Style.NORMAL + self.summary + Fore.WHITE + 
              " Created: " + str(self.create_date.month) + "/" + str(self.create_date.day) + "/" + str(self.create_date.year) +
              "\n    " + Fore.BLUE + Style.BRIGHT + self.team + Fore.RED + Style.NORMAL + " Estimate: " + str(self.estimate) +
              "\n    " + Fore.YELLOW + Style.NORMAL + f"{self.issues_with_points} issues have points and {self.issues_with_no_points} don't. {self.issues_points} points total." +
              "\n    " + Fore.MAGENTA + " Sub Labels: " + Fore.WHITE + str(self.sub_labels) + Style.RESET_ALL)
        
    def excel_worksheet_create(ws, epics, jira_issue_link, project_figma_link, project_label):
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 40

        table = Table(displayName="TableEpics", ref="A1:H" + str(len(epics) + 1))
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style

        # Populate data
        ws.append(["Epic", "Summary", "Team", "Estimate", "Issues w Points", "Issues w No Points ", "Issues Total Points", "Sub Labels"])
        for epicitem in epics:
            sub_labels = ""
            coun_label = 0
            for label in epicitem.sub_labels:
                if label != project_label:
                    if coun_label == 0:
                        sub_labels += label
                        coun_label += 1
                    else:
                        sub_labels += ", " + label
            ws.append([epicitem.key, epicitem.summary, epicitem.team, epicitem.estimate, epicitem.issues_with_points, epicitem.issues_with_no_points, epicitem.issues_points, sub_labels])    

        ws.add_table(table)

        # Format Data
        for row in ws[2:ws.max_row]:  # Exclude The Header
            cell = row[0] # zeor based index
            value_use = cell.value
            cell.hyperlink = f"{jira_issue_link}{value_use}"
            cell.value = value_use
            cell.style = "Hyperlink"
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[0] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[2:ws.max_row]:  # skip the header
            cell = row[1] # zeor based index
            cell.alignment = Alignment(wrap_text=True)
            cell.number_format = "text"

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[2] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[3] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[4] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[5] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[6] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add Figma Plan Link to the bottom of the Epics Sheet
        ws["A" + str(len(epics) + 3)].hyperlink = project_figma_link
        ws["A" + str(len(epics) + 3)].value = "Figma Plan"
        ws["A" + str(len(epics) + 3)].style = "Hyperlink"

    def excel_worksheet_summary(ws, epics, project_label, project_created):
        def test_zero_value(value, cell):
            if value == 0:
                cell.value = " - "
            else:
                cell.value = value

        sub_labels = []

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 25

        epic_total = 0
        epic_estimate_total = 0
        epic_with_estimate = 0
        epic_estimate_max = 0
        epic_estimate_min = 0
        epic_estimate_avg = 0
        epic_percent_with_estimate = 0
        issue_total = 0
        issue_estimate_total = 0
        issue_with_estimate = 0
        issue_estimate_max = 0
        issue_estimate_min = 0
        issue_estimate_avg = 0
        issue_percent_with_estimate = 0

        for epicitem in epics:
            # List of all sub labels
            for label in epicitem.sub_labels:
                if label not in sub_labels:
                    if label != project_label:
                        sub_labels.append(label)
            # Epic Summary Data 
            epic_total += 1
            if epicitem.estimate != None:
                epic_with_estimate += 1
                epic_estimate_total += epicitem.estimate
                if epicitem.estimate > epic_estimate_max:
                    epic_estimate_max = epicitem.estimate
                if epicitem.estimate < epic_estimate_min:
                    epic_estimate_min = epicitem.estimate
            for issueitem in epicitem.issues:
                issue_total += 1
                if issueitem.size != None:
                    issue_with_estimate += 1
                    issue_estimate_total += issueitem.size
                    if issueitem.size > issue_estimate_max:
                        issue_estimate_max = issueitem.size
                    if issueitem.size < issue_estimate_min:
                        issue_estimate_min = issueitem.size
        if epic_estimate_total > 0:
            epic_estimate_avg = epic_estimate_total / epic_with_estimate
        else:
            epic_estimate_avg = 0
        
        if epic_with_estimate > 0:
            epic_percent_with_estimate = epic_with_estimate / epic_total
        else:
            epic_percent_with_estimate = 0

        if issue_estimate_total > 0:
            issue_estimate_avg = issue_estimate_total / issue_with_estimate
        else:
            issue_estimate_avg = 0
        
        if issue_with_estimate > 0:
            issue_percent_with_estimate = issue_with_estimate / issue_total
        else:
            issue_percent_with_estimate = 0

        ws["E3"] = "Project Label"
        ws["E3"].font = Font(bold=True, size=14)
        ws["E4"] = project_label
        ws["E4"].font = Font(italic=True, size=12)
        ws["A1"] = "Created"
        ws["A1"].font = Font(bold=True, size=14)
        ws["B1"] = project_created
        ws["B1"].font = Font(italic=True, size=12)

        ws["A3"] = "Epics"
        ws["A3"].font = Font(bold=True, size=14)
        ws["C3"] = "All Issues"
        ws["C3"].font = Font(bold=True, size=14)

        ws["A4"] = "Count"
        ws["A5"] = "Total Estimate"
        ws["A6"] = "With Estimates"
        ws["A7"] = "Percent with Est"
        ws["A8"] = "Average Estimate"
        ws["A9"] = "Max Estimate"
        ws["A10"] = "Min Estimate"

        test_zero_value(epic_total, ws["B4"])
        test_zero_value(epic_estimate_total, ws["B5"])
        test_zero_value(epic_with_estimate, ws["B6"])
        test_zero_value(epic_percent_with_estimate, ws["B7"])
        ws["B7"].number_format = numbers.FORMAT_PERCENTAGE_00
        test_zero_value(epic_estimate_avg, ws["B8"])
        test_zero_value(epic_estimate_max, ws["B9"])
        test_zero_value(epic_estimate_min, ws["B10"])
        
        for row in ws[4:ws.max_row]:  # 1 Based Index
            cell = row[1] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws["C4"] = "Count"
        ws["C5"] = "Total Estimate"
        ws["C6"] = "With Estimates"
        ws["C7"] = "Percent with Est"
        ws["C8"] = "Average Estimate"
        ws["C9"] = "Max Estimate"
        ws["C10"] = "Min Estimate"

        test_zero_value(issue_total, ws["D4"])
        test_zero_value(issue_estimate_total, ws["D5"])
        test_zero_value(issue_with_estimate, ws["D6"])
        test_zero_value(issue_percent_with_estimate, ws["D7"])
        ws["D7"].number_format = numbers.FORMAT_PERCENTAGE_00
        test_zero_value(issue_estimate_avg, ws["D8"])
        test_zero_value(issue_estimate_max, ws["D9"])
        test_zero_value(issue_estimate_min, ws["D10"])

        for row in ws[4:ws.max_row]:  # 1 Based Index
            cell = row[3] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws["E5"] = "Sub Labels"
        ws["E5"].font = Font(bold=True, size=14)
        
        start_sub = 6
        for Label in sub_labels:
            ws["E" + str(start_sub)] = Label
            start_sub += 1

        for row in ws[6:ws.max_row]:  # 1 Based Index
            cell = row[4] # zeor based index
            cell.alignment = Alignment(horizontal="left", vertical="center")