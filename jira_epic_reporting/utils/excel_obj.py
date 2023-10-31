from typing import List
from objects.issue import Issue
from colorama import init, Fore, Back, Style
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, numbers, PatternFill, Border, Side, colors, GradientFill
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice, GradientFillProperties
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors
from copy import deepcopy
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor
from openpyxl.styles.fills import Stop
import console_util
import utils.claude_util as claude_util
from utils.jira_obj import Jira
from objects.project import Project
from objects.epic import Epic
from objects.sprint import Sprint
from objects.user import User
from objects.board import Board
from objects.status import Status
from objects.issue_type import IssueType
from openpyxl.chart import ScatterChart, Reference, Series, BarChart

class Excel:
    def __init__(self, claudekey, project_label,jira_issue_link, create_date, ai_out, other_links):
        self.claudekey = claudekey
        self.project_label = project_label
        self.jira_issue_link = jira_issue_link
        self.create_date = create_date
        self.ai_out = ai_out
        self.other_links = other_links

    def get_project_sub_labels(self, epics, project_label):
        sub_labels = []
        for epicitem in epics:
            for label in epicitem.sub_labels:
                if label not in sub_labels:
                    if label != project_label:
                        sub_labels.append(label)
        return sub_labels

    def get_epics_with_sub_label(self, epics, sub_label):
        epics_with_sub_label = []
        for epicitem in epics:
            for label in epicitem.sub_labels:
                if label == sub_label:
                    # Check for issues in Epic and don't add if no issues
                    if len(epicitem.issues) > 0:
                        epics_with_sub_label.append(epicitem)
        return epics_with_sub_label

    def excel_worksheet_summary(self, ws, epics, chart, chart_stories):

        sub_labels = self.get_project_sub_labels(epics, self.project_label)

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 25

        epic_total = 0
        epic_estimate_total = 0
        epic_with_estimate = 0
        epic_estimate_max = 0
        epic_estimate_min = 0
        epic_estimate_avg = 0
        epic_percent_with_estimate = 0
        issue_total = 0
        issue_estimate_total = 0
        issue_with_estimate = 0
        issue_estimate_max = 0
        issue_estimate_min = 0
        issue_estimate_avg = 0
        issue_percent_with_estimate = 0

        for epicitem in epics:
            # Epic Summary Data 
            epic_total += 1
            if epicitem.estimate != None:
                epic_with_estimate += 1
                epic_estimate_total += epicitem.estimate
                if epicitem.estimate > epic_estimate_max:
                    epic_estimate_max = epicitem.estimate
                if epicitem.estimate < epic_estimate_min or epic_estimate_min == 0:
                    epic_estimate_min = epicitem.estimate
            for issueitem in epicitem.issues:
                issue_total += 1
                if issueitem.size != None:
                    issue_with_estimate += 1
                    issue_estimate_total += issueitem.size
                    if issueitem.size > issue_estimate_max:
                        issue_estimate_max = issueitem.size
                    if issueitem.size < issue_estimate_min or issue_estimate_min == 0:
                        issue_estimate_min = issueitem.size
        if epic_estimate_total > 0:
            epic_estimate_avg = epic_estimate_total / epic_with_estimate
        else:
            epic_estimate_avg = 0
        
        if epic_with_estimate > 0:
            epic_percent_with_estimate = epic_with_estimate / epic_total
        else:
            epic_percent_with_estimate = 0

        if issue_estimate_total > 0:
            issue_estimate_avg = issue_estimate_total / issue_with_estimate
        else:
            issue_estimate_avg = 0
        
        if issue_with_estimate > 0:
            issue_percent_with_estimate = issue_with_estimate / issue_total
        else:
            issue_percent_with_estimate = 0

        ws["E3"] = "Project Label"
        ws["E3"].font = Font(bold=True, size=14)
        ws["E4"] = self.project_label
        ws["E4"].font = Font(italic=True, size=12)
        ws["A1"] = "Created"
        ws["A1"].font = Font(bold=True, size=14)
        ws["B1"] = self.create_date
        ws["B1"].font = Font(italic=True, size=12)

        ws["A3"].value = '=HYPERLINK("#Epics!A1","Epics")'
        ws["A3"].style = "Hyperlink"
        ws["A3"].font = Font(bold=True, underline="single", size=14)

        ws["C3"] = '=HYPERLINK("#Issues!A1","All Issues")'
        ws["C3"].style = "Hyperlink"
        ws["C3"].font = Font(bold=True, underline="single", size=14)

        ws["A4"] = "Count"
        ws["A5"] = "Total Estimate"
        ws["A6"] = "With Estimates"
        ws["A7"] = "Percent with Est"
        ws["A8"] = "Average Estimate"
        ws["A9"] = "Max Estimate"
        ws["A10"] = "Min Estimate"

        Jira.test_zero_value(epic_total, ws["B4"])
        Jira.test_zero_value(epic_estimate_total, ws["B5"])
        Jira.test_zero_value(epic_with_estimate, ws["B6"])
        Jira.test_zero_value(epic_percent_with_estimate, ws["B7"])
        ws["B7"].number_format = numbers.FORMAT_PERCENTAGE_00
        Jira.test_zero_value(epic_estimate_avg, ws["B8"])
        Jira.test_zero_value(epic_estimate_max, ws["B9"])
        Jira.test_zero_value(epic_estimate_min, ws["B10"])
        
        for row in ws[4:ws.max_row]:  # 1 Based Index
            cell = row[1] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        link_location = 12
        if len(self.other_links) > 0:
            for key, value in self.other_links.items():
                ws["A" + str(link_location)].hyperlink = value
                ws["A" + str(link_location)].value = key
                ws["A" + str(link_location)].style = "Hyperlink"
                ws["A" + str(link_location)].font = Font(bold=True, size=14)
                link_location += 1

        ws["C4"] = "Count"
        ws["C5"] = "Total Estimate"
        ws["C6"] = "With Estimates"
        ws["C7"] = "Percent with Est"
        ws["C8"] = "Average Estimate"
        ws["C9"] = "Max Estimate"
        ws["C10"] = "Min Estimate"

        Jira.test_zero_value(issue_total, ws["D4"])
        Jira.test_zero_value(issue_estimate_total, ws["D5"])
        Jira.test_zero_value(issue_with_estimate, ws["D6"])
        Jira.test_zero_value(issue_percent_with_estimate, ws["D7"])
        ws["D7"].number_format = numbers.FORMAT_PERCENTAGE_00
        Jira.test_zero_value(issue_estimate_avg, ws["D8"])
        Jira.test_zero_value(issue_estimate_max, ws["D9"])
        Jira.test_zero_value(issue_estimate_min, ws["D10"])

        for row in ws[4:ws.max_row]:  # 1 Based Index
            cell = row[3] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws["E5"] = "Sub Labels"
        ws["E5"].font = Font(bold=True, size=14)
        
        start_sub = 6
        for Label in sub_labels:
            ws["E" + str(start_sub)] = Label
            start_sub += 1

        for row in ws[6:ws.max_row]:  # 1 Based Index
            cell = row[4] # zeor based index
            cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.add_chart(chart, "A12")
        anchor = TwoCellAnchor()
        anchor._from.col = 0 # A
        anchor._from.row = 16 
        anchor.to.col = 10 # 
        anchor.to.row = 40 # row 
        chart.anchor = anchor

        ws.add_chart(chart_stories, "A42")
        anchor = TwoCellAnchor()
        anchor._from.col = 0 # A
        anchor._from.row = 42 
        anchor.to.col = 10 # 
        anchor.to.row = 62 # row 
        chart_stories.anchor = anchor        

    def excel_worksheet_create_epics(self, ws, epics):

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 40

        table = Table(displayName="TableEpics", ref="A1:H" + str(len(epics) + 1))
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style

        # Populate data
        ws.append(["Epic", "Summary", "Team", "Estimate", "Issues w Points", "Issues w No Points ", "Issues Total Points", "Sub Labels"])
        for epicitem in epics:
            sub_labels = ""
            coun_label = 0
            for label in epicitem.sub_labels:
                if label != self.project_label:
                    if coun_label == 0:
                        sub_labels += label
                        coun_label += 1
                    else:
                        sub_labels += ", " + label
            ws.append([epicitem.key, epicitem.summary, epicitem.team, epicitem.estimate, epicitem.issues_with_points, epicitem.issues_with_no_points, epicitem.issues_points, sub_labels])    

        ws.add_table(table)

        # Format Data
        
        for row in ws[1:ws.max_row]:  # Exclude The Header
            cell = row[0] # zeor based index
            value_use = cell.value
            cell.hyperlink = f"{self.jira_issue_link}{value_use}"
            cell.value = value_use
            cell.style = "Hyperlink"
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[0] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # skip the header
            cell = row[1] # zeor based index
            cell.alignment = Alignment(wrap_text=True)
            cell.number_format = "text"
            
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[2] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[3] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[4] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[5] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[6] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def excel_worksheet_story_add(self, ws, all_sprints):
        """
            Create the Story Add Chart
        """
        # First create the data to use
        ws.column_dimensions["A"].width = 18 # Sprint Number
        ws.column_dimensions["B"].width = 18 # Total Stories


        table = Table(displayName="TableStories", ref="A1:B" + str(len(all_sprints) + 1))
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style

        # Populate data
        ws.append(["Sprint", "Stories Added"])
        for sprintitem in all_sprints:
            ws.append([sprintitem.name, sprintitem.stories_created])
        ws.add_table(table)

        chart = BarChart()
        chart.title = "Stories Added By Sprint"
        chart.type = "col"
        chart.style = 10
        chart.x_axis.title = 'Sprints'
        chart.y_axis.title = 'Count'
        max_row_set = ws.max_row
        
        data = Reference(ws, min_col=2, min_row=1, max_row=max_row_set, max_col=2)
        cats = Reference(ws, min_col=1, min_row=2, max_row=max_row_set)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4 

        chart2 = deepcopy(chart)

        ws.add_chart(chart, "F2")
        anchor = TwoCellAnchor()
        anchor._from.col = 5 # F
        anchor._from.row = 2 # row 19, using 0-based indexing
        anchor.to.col = 20 # 
        anchor.to.row = 25 # row 
        chart.anchor = anchor        

        return chart2

    def excel_worksheet_burnup(self, ws, all_sprints):
        """
            Create the burnup chart
        """
        # First create the data to use
        ws.column_dimensions["A"].width = 18 # Sprint Number
        ws.column_dimensions["B"].width = 18 # Sprint Created Accum Points
        ws.column_dimensions["C"].width = 18 # Sprint Completed Accum Points
        ws.column_dimensions["D"].width = 18 # Trajectory

        table = Table(displayName="TableBurnup", ref="A1:D" + str(len(all_sprints) + 1))
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style

        # Populate data
        ws.append(["Sprint", "Accum Pointed", "Accum Completed", "Trajectory"])
        accum_points = 0
        accum_completed = 0
        location = 0
        for sprintitem in all_sprints:
            accum_points += sprintitem.story_points_created
            accum_completed += sprintitem.story_points_completed
            if location == 0:
                ws.append([sprintitem.name, accum_points, accum_completed, 0])
            elif (location +1) == len(all_sprints):
                ws.append([sprintitem.name, accum_points, accum_completed, accum_points])
            else:
                ws.append([sprintitem.name, accum_points, accum_completed, "=NA()"])
            location += 1
        ws.add_table(table)

        chart = ScatterChart()
        chart.title = "Burn Up Chart"
        chart.style = 6
        chart.x_axis.title = 'Sprints'
        chart.y_axis.title = 'Points'
        max_row_set = ws.max_row
        xvalues = Reference(ws, min_col=1, min_row=2, max_row=max_row_set) # Size Column

        for i in range(2, 5): # Column (2, 5) 2,3,4: B and C and D
            values = Reference(ws, min_col=i, min_row=1, max_row=max_row_set)
            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series) 

        # Chart Line Properties
        series2 = chart.series[2]
        line_props = LineProperties(solidFill="ABFA96", prstDash="dash")
        series2.graphicalProperties.line = line_props

        series0 = chart.series[0]
        line_props0 = LineProperties(solidFill="04A24E", prstDash="solid")
        series0.graphicalProperties.line = line_props0

        series1 = chart.series[1]
        line_props1 = LineProperties(solidFill="FF0000", prstDash="dot")
        series1.graphicalProperties.line = line_props1

        chart2 = deepcopy(chart)

        ws.add_chart(chart, "F2")
        anchor = TwoCellAnchor()
        anchor._from.col = 5 # F
        anchor._from.row = 2 # row 19, using 0-based indexing
        anchor.to.col = 20 # 
        anchor.to.row = 48 # row 
        chart.anchor = anchor

        return chart2

    def excel_worksheet_create_issues(self, ws, epics, table_name):
        def set_date(value):
            if value != "":
                return value.strftime("%m/%d/%Y")
            else:
                return ""

        # Start Building Issues Tab
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 40
        ws.column_dimensions["I"].width = 20
        ws.column_dimensions["J"].width = 20
        ws.column_dimensions["K"].width = 20
        ws.column_dimensions["L"].width = 20
        ws.column_dimensions["M"].width = 30
        ws.column_dimensions["N"].width = 30
        ws.column_dimensions["O"].width = 30
        ws.column_dimensions["P"].width = 30



        all_issues = 0
        for epicitem in epics:
            all_issues += (len(epicitem.issues))
        table_issues = Table(displayName=table_name, ref="A1:P" + str(all_issues + 1))
        style_issues = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table_issues.tableStyleInfo = style_issues

        ws.append(["Issue", "Summary", "Team", "Size", 
                    "Status", "Priority", "Type", 
                    "Project", "Assignee", "Created", 
                    "Updated", "Epic",
                    "Sprint", "Sprint 2", "Sprint 3", "Sprint 4"])
        for epicitem in epics:
            for issueitem in epicitem.issues:
                sprint_count = 0
                sprint1 = ""
                sprint2 = ""
                sprint3 = ""
                sprint4 = ""
                for sprintitem in issueitem.sprint:
                    if sprint_count == 0:
                        sprint1 = sprintitem.name
                    elif sprint_count == 1:
                        sprint2 = sprintitem.name
                    elif sprint_count == 2:
                        sprint3 = sprintitem.name
                    elif sprint_count == 3:
                        sprint4 = sprintitem.name
                    sprint_count += 1

                ws.append([issueitem.key, issueitem.summary, epicitem.team, issueitem.size
                            , issueitem.status, issueitem.priority, issueitem.issuetype
                            , issueitem.project_name, issueitem.assignee_displayName, set_date(issueitem.created)
                            , set_date(issueitem.updated), epicitem.key,
                            sprint1, sprint2, sprint3, sprint4])

        ws.add_table(table_issues)

        for row in ws[1:ws.max_row]:  # Exclude The Header
            cell = row[0] # zeor based index
            value_use = cell.value
            cell.hyperlink = f"{self.jira_issue_link}{value_use}"
            cell.value = value_use
            cell.style = "Hyperlink"
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[0] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[2] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[3] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[4] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[5] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[6] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[7] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[8] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[9] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[10] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Exclude The Header
            cell = row[11] # zeor based index
            value_use = cell.value
            cell.hyperlink = f"{self.jira_issue_link}{value_use}"
            cell.value = value_use
            cell.style = "Hyperlink"
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[11] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def create_label_excel_report(self, epics, all_sprints):
        console_util.terminal_update("Creating Excel Document", " - ", False)
        workbook = Workbook()
        worksheet_summary = workbook.active
        worksheet_summary.title = "Summary"
        worksheet_burnup = workbook.create_sheet("Burnup")
        worksheet_story_add = workbook.create_sheet("Story Add")
        worksheet_epics = workbook.create_sheet("All Epics")
        worksheet_issues = workbook.create_sheet("All Issues")
        
        # Create the burnup chart worksheet
        chart_copy = self.excel_worksheet_burnup(worksheet_burnup, all_sprints)

        # Create Story Add Chart
        chart_stories = self.excel_worksheet_story_add(worksheet_story_add, all_sprints)
        
        # Create the Summary Tab
        self.excel_worksheet_summary(worksheet_summary, epics, chart_copy, chart_stories)

        # Create the Epic Tab
        self.excel_worksheet_create_epics(worksheet_epics, epics)

        # Create the Issue Tab
        self.excel_worksheet_create_issues(worksheet_issues, epics, "TableAllIssues")

        # Create the Sub Label Tabs
        sub_labels = self.get_project_sub_labels(epics, self.project_label)
        for sub_label in sub_labels:
            sheet_epics = self.get_epics_with_sub_label(epics, sub_label)
            if len(sheet_epics) > 0:
                worksheet_sub_label = workbook.create_sheet(sub_label)
                self.excel_worksheet_create_issues(worksheet_sub_label, sheet_epics, f"Table{sub_label}Issues")

        if self.ai_out:
            worksheet_ai = workbook.create_sheet("Epic Health")
            claude_util.excel_worksheet_ai_create(worksheet_ai, epics, self.jira_issue_link, self.claudekey)   

        return workbook

    def create_jira_info_report(self, boards, sprints, users, projects):
        console_util.terminal_update("Creating Jira Excel Document", " - ", False)
        workbook = Workbook()
        worksheet_boards = workbook.active
        worksheet_boards.title = "Boards"
        worksheet_sprints = workbook.create_sheet("All Sprints")
        worksheet_statuses = workbook.create_sheet("All Statuses")
        worksheet_users = workbook.create_sheet("Users")
        self.excel_boards(worksheet_boards, boards)
        self.excel_sprints(worksheet_sprints, sprints)
        self.excel_statuses(worksheet_statuses, projects)
        self.excel_users(worksheet_users, users)    
        return workbook

    def excel_boards(self, ws, boards):
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 25
        
        table = Table(displayName="TableBoards", ref="A1:C" + str(len(boards) + 1))
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style

        ws.append(["Board ID", "Board Name", "Board Type"])
        for board in boards:
            ws.append([board.id,board.name,board.type])

        ws.add_table(table)

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[0] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row in ws[2:ws.max_row]:  # Include The Header
            cell = row[1] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[3:ws.max_row]:  # Include The Header
            cell = row[1] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def excel_sprints(self, ws, sprints):
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 15

        ws.append(["Sprint ID", "Board ID", "Name", "State"])
        
        table = Table(displayName="TableSprints", ref="A1:D" + str(len(sprints) + 1))
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style

        for sprint in sprints:
            ws.append([sprint.id, sprint.boardID, sprint.name, sprint.state])

        ws.add_table(table)
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[0] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[1] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[2] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[3] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def excel_statuses(self, ws, projects):
        """
        List all statuses with the projects
        """
        ws.column_dimensions["A"].width = 18 # Project ID
        ws.column_dimensions["B"].width = 15 # Project Key
        ws.column_dimensions["C"].width = 30 # Project Name
        ws.column_dimensions["D"].width = 18 # Type ID
        ws.column_dimensions["E"].width = 30 # Type Name
        ws.column_dimensions["F"].width = 18 # Status ID
        ws.column_dimensions["G"].width = 35 # Status Name
        ws.column_dimensions["H"].width = 70 # Status Description

        ws.append(["Proj ID", "Project Key", "Proj Name", "Type ID", "Type Name", "Status ID", "Status Name", "Status Description"])

        total_length = 1
        for project in projects:
            for issue in project.issues:
                for status in issue.statuses:
                    total_length += 1

        table = Table(displayName="TableStatuses", ref="A1:H" + str(total_length))
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style

        for project in projects:
            for issue in project.issues:
                for status in issue.statuses:
                    ws.append([project.id, project.key, project.name, issue.id, issue.name, status.id, status.name, status.description])

        ws.add_table(table)

    def excel_users(self, ws, users):
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 70

        ws.append(["Name", "Active", "email", "ID"])

        table = Table(displayName="TableUsers", ref="A1:D" + str(len(users) + 1))
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style

        for user in users:
            ws.append([user.name, user.active, user.email, user.id])

        ws.add_table(table)

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[0] # zeor based index
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[1] # zeor based index
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[2] # zeor based index
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        for row in ws[1:ws.max_row]:  # Include The Header
            cell = row[3] # zeor based index
            cell.alignment = Alignment(horizontal="left", vertical="center")