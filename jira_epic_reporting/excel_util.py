
from typing import List
from issue import Issue
from colorama import init, Fore, Back, Style
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter, quote_sheetname
import jira_utils

def get_project_sub_labels(epics, project_label):
    sub_labels = []
    for epicitem in epics:
        for label in epicitem.sub_labels:
            if label not in sub_labels:
                if label != project_label:
                    sub_labels.append(label)
    return sub_labels

def get_epics_with_sub_label(epics, sub_label):
    epics_with_sub_label = []
    for epicitem in epics:
        for label in epicitem.sub_labels:
            if label == sub_label:
                # Check for issues in Epic and don't add if no issues
                if len(epicitem.issues) > 0:
                    epics_with_sub_label.append(epicitem)
    return epics_with_sub_label

def excel_worksheet_summary(ws, epics, project_label, project_created, other_links):

    sub_labels = get_project_sub_labels(epics, project_label)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 25

    epic_total = 0
    epic_estimate_total = 0
    epic_with_estimate = 0
    epic_estimate_max = 0
    epic_estimate_min = 0
    epic_estimate_avg = 0
    epic_percent_with_estimate = 0
    issue_total = 0
    issue_estimate_total = 0
    issue_with_estimate = 0
    issue_estimate_max = 0
    issue_estimate_min = 0
    issue_estimate_avg = 0
    issue_percent_with_estimate = 0

    for epicitem in epics:
        # Epic Summary Data 
        epic_total += 1
        if epicitem.estimate != None:
            epic_with_estimate += 1
            epic_estimate_total += epicitem.estimate
            if epicitem.estimate > epic_estimate_max:
                epic_estimate_max = epicitem.estimate
            if epicitem.estimate < epic_estimate_min:
                epic_estimate_min = epicitem.estimate
        for issueitem in epicitem.issues:
            issue_total += 1
            if issueitem.size != None:
                issue_with_estimate += 1
                issue_estimate_total += issueitem.size
                if issueitem.size > issue_estimate_max:
                    issue_estimate_max = issueitem.size
                if issueitem.size < issue_estimate_min:
                    issue_estimate_min = issueitem.size
    if epic_estimate_total > 0:
        epic_estimate_avg = epic_estimate_total / epic_with_estimate
    else:
        epic_estimate_avg = 0
    
    if epic_with_estimate > 0:
        epic_percent_with_estimate = epic_with_estimate / epic_total
    else:
        epic_percent_with_estimate = 0

    if issue_estimate_total > 0:
        issue_estimate_avg = issue_estimate_total / issue_with_estimate
    else:
        issue_estimate_avg = 0
    
    if issue_with_estimate > 0:
        issue_percent_with_estimate = issue_with_estimate / issue_total
    else:
        issue_percent_with_estimate = 0

    ws["E3"] = "Project Label"
    ws["E3"].font = Font(bold=True, size=14)
    ws["E4"] = project_label
    ws["E4"].font = Font(italic=True, size=12)
    ws["A1"] = "Created"
    ws["A1"].font = Font(bold=True, size=14)
    ws["B1"] = project_created
    ws["B1"].font = Font(italic=True, size=12)

    ws["A3"].value = '=HYPERLINK("#Epics!A1","Epics")'
    ws["A3"].style = "Hyperlink"
    ws["A3"].font = Font(bold=True, underline="single", size=14)

    ws["C3"] = '=HYPERLINK("#Issues!A1","All Issues")'
    ws["C3"].style = "Hyperlink"
    ws["C3"].font = Font(bold=True, underline="single", size=14)

    ws["A4"] = "Count"
    ws["A5"] = "Total Estimate"
    ws["A6"] = "With Estimates"
    ws["A7"] = "Percent with Est"
    ws["A8"] = "Average Estimate"
    ws["A9"] = "Max Estimate"
    ws["A10"] = "Min Estimate"

    jira_utils.test_zero_value(epic_total, ws["B4"])
    jira_utils.test_zero_value(epic_estimate_total, ws["B5"])
    jira_utils.test_zero_value(epic_with_estimate, ws["B6"])
    jira_utils.test_zero_value(epic_percent_with_estimate, ws["B7"])
    ws["B7"].number_format = numbers.FORMAT_PERCENTAGE_00
    jira_utils.test_zero_value(epic_estimate_avg, ws["B8"])
    jira_utils.test_zero_value(epic_estimate_max, ws["B9"])
    jira_utils.test_zero_value(epic_estimate_min, ws["B10"])
    
    for row in ws[4:ws.max_row]:  # 1 Based Index
        cell = row[1] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    link_location = 12
    if len(other_links) > 0:
        for key, value in other_links.items():
            ws["A" + str(link_location)].hyperlink = value
            ws["A" + str(link_location)].value = key
            ws["A" + str(link_location)].style = "Hyperlink"
            ws["A" + str(link_location)].font = Font(bold=True, size=14)
            link_location += 1

    ws["C4"] = "Count"
    ws["C5"] = "Total Estimate"
    ws["C6"] = "With Estimates"
    ws["C7"] = "Percent with Est"
    ws["C8"] = "Average Estimate"
    ws["C9"] = "Max Estimate"
    ws["C10"] = "Min Estimate"

    jira_utils.test_zero_value(issue_total, ws["D4"])
    jira_utils.test_zero_value(issue_estimate_total, ws["D5"])
    jira_utils.test_zero_value(issue_with_estimate, ws["D6"])
    jira_utils.test_zero_value(issue_percent_with_estimate, ws["D7"])
    ws["D7"].number_format = numbers.FORMAT_PERCENTAGE_00
    jira_utils.test_zero_value(issue_estimate_avg, ws["D8"])
    jira_utils.test_zero_value(issue_estimate_max, ws["D9"])
    jira_utils.test_zero_value(issue_estimate_min, ws["D10"])

    for row in ws[4:ws.max_row]:  # 1 Based Index
        cell = row[3] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws["E5"] = "Sub Labels"
    ws["E5"].font = Font(bold=True, size=14)
    
    start_sub = 6
    for Label in sub_labels:
        ws["E" + str(start_sub)] = Label
        start_sub += 1

    for row in ws[6:ws.max_row]:  # 1 Based Index
        cell = row[4] # zeor based index
        cell.alignment = Alignment(horizontal="left", vertical="center")

def excel_worksheet_create_epics(ws, epics, jira_issue_link, project_label):

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 40

    table = Table(displayName="TableEpics", ref="A1:H" + str(len(epics) + 1))
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style

    # Populate data
    ws.append(["Epic", "Summary", "Team", "Estimate", "Issues w Points", "Issues w No Points ", "Issues Total Points", "Sub Labels"])
    for epicitem in epics:
        sub_labels = ""
        coun_label = 0
        for label in epicitem.sub_labels:
            if label != project_label:
                if coun_label == 0:
                    sub_labels += label
                    coun_label += 1
                else:
                    sub_labels += ", " + label
        ws.append([epicitem.key, epicitem.summary, epicitem.team, epicitem.estimate, epicitem.issues_with_points, epicitem.issues_with_no_points, epicitem.issues_points, sub_labels])    

    ws.add_table(table)

    # Format Data
    for row in ws[2:ws.max_row]:  # Exclude The Header
        cell = row[0] # zeor based index
        value_use = cell.value
        cell.hyperlink = f"{jira_issue_link}{value_use}"
        cell.value = value_use
        cell.style = "Hyperlink"
    
    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[0] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[2:ws.max_row]:  # skip the header
        cell = row[1] # zeor based index
        cell.alignment = Alignment(wrap_text=True)
        cell.number_format = "text"

    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[2] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[3] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[4] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[5] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[6] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

def excel_worksheet_create_issues(ws, epics, jira_issue_link, table_name):
    def set_date(value):
        if value != "":
            return value.strftime("%m/%d/%Y")
        else:
            return ""

    # Start Building Issues Tab
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 40
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 20
    ws.column_dimensions["K"].width = 20
    ws.column_dimensions["L"].width = 20
    ws.column_dimensions["M"].width = 30
    ws.column_dimensions["N"].width = 30
    ws.column_dimensions["O"].width = 30
    ws.column_dimensions["P"].width = 30



    all_issues = 0
    for epicitem in epics:
        all_issues += (len(epicitem.issues))
    table_issues = Table(displayName=table_name, ref="A1:P" + str(all_issues + 1))
    style_issues = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table_issues.tableStyleInfo = style_issues

    ws.append(["Issue", "Summary", "Team", "Size", 
                "Status", "Priority", "Type", 
                "Project", "Assignee", "Created", 
                "Updated", "Epic",
                "Sprint", "Sprint 2", "Sprint 3", "Sprint 4"])
    for epicitem in epics:
        for issueitem in epicitem.issues:
            sprint_count = 0
            sprint1 = ""
            sprint2 = ""
            sprint3 = ""
            sprint4 = ""
            for sprintitem in issueitem.sprint:
                if sprint_count == 0:
                    sprint1 = sprintitem.name
                elif sprint_count == 1:
                    sprint2 = sprintitem.name
                elif sprint_count == 2:
                    sprint3 = sprintitem.name
                elif sprint_count == 3:
                    sprint4 = sprintitem.name
                sprint_count += 1

            ws.append([issueitem.key, issueitem.summary, epicitem.team, issueitem.size
                        , issueitem.status, issueitem.priority, issueitem.issuetype
                        , issueitem.project_name, issueitem.assignee_displayName, set_date(issueitem.created)
                        , set_date(issueitem.updated), epicitem.key,
                        sprint1, sprint2, sprint3, sprint4])

    ws.add_table(table_issues)

    for row in ws[2:ws.max_row]:  # Exclude The Header
        cell = row[0] # zeor based index
        value_use = cell.value
        cell.hyperlink = f"{jira_issue_link}{value_use}"
        cell.value = value_use
        cell.style = "Hyperlink"
    
    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[0] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[2] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[3] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[4] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[5] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[6] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[7] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[8] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[9] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[10] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws[2:ws.max_row]:  # Exclude The Header
        cell = row[11] # zeor based index
        value_use = cell.value
        cell.hyperlink = f"{jira_issue_link}{value_use}"
        cell.value = value_use
        cell.style = "Hyperlink"
    
    for row in ws[1:ws.max_row]:  # Include The Header
        cell = row[11] # zeor based index
        cell.alignment = Alignment(horizontal="center", vertical="center")